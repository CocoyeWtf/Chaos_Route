"""Tests export Excel de l'historique des tours (ticket #17, phase 1).

Vérifie que /api/exports/tour-history renvoie un classeur Excel reprenant les
colonnes complètes de la page « Historique des tours » : une ligne par tour,
libellés FR (nature, statut), véhicule/transporteur résolus depuis le contrat,
nombre d'arrêts, et respect du filtre région.
"""

import io
import uuid

import pytest
from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "Code", "Nature", "Date", "Base", "Véhicule", "Transporteur", "Arrêts",
    "EQC", "Km", "Coût (€)", "Statut", "Départ prévu", "Priorité",
    "Retour prévu", "Chauffeur", "Arrivée chauffeur", "Fin chargement",
    "Top départ", "Sortie barrière", "Retour barrière",
]


async def _make_base(db_session, region, name="Base Hist"):
    from app.models.base_logistics import BaseLogistics

    base = BaseLogistics(
        code=f"B{uuid.uuid4().hex[:5].upper()}", name=name, region_id=region.id,
    )
    db_session.add(base)
    await db_session.commit()
    await db_session.refresh(base)
    return base


async def _make_contract(db_session, region):
    from app.models.contract import Contract

    contract = Contract(
        transporter_name="Transports Test",
        code=f"C{uuid.uuid4().hex[:5].upper()}",
        region_id=region.id,
        vehicle_code=f"V{uuid.uuid4().hex[:4].upper()}",
        vehicle_name="Semi frigo",
    )
    db_session.add(contract)
    await db_session.commit()
    await db_session.refresh(contract)
    return contract


async def _make_tour(db_session, base, *, code, date, contract=None, n_stops=0, region=None, **kwargs):
    from app.models.tour import Tour, TourStatus
    from app.models.tour_stop import TourStop
    from app.models.pdv import PDV, PDVType

    tour = Tour(
        date=date, code=code, base_id=base.id,
        status=kwargs.pop("status", TourStatus.COMPLETED),
        contract_id=contract.id if contract else None,
        **kwargs,
    )
    db_session.add(tour)
    await db_session.commit()
    await db_session.refresh(tour)

    for seq in range(1, n_stops + 1):
        pdv = PDV(
            code=f"P{uuid.uuid4().hex[:5].upper()}", name="PDV",
            type=PDVType.HYPER, region_id=region.id,
        )
        db_session.add(pdv)
        await db_session.commit()
        await db_session.refresh(pdv)
        db_session.add(TourStop(tour_id=tour.id, pdv_id=pdv.id, sequence_order=seq, eqp_count=1))
    await db_session.commit()
    return tour


@pytest.mark.asyncio
async def test_tour_history_export_columns_and_values(client, db_session, test_region):
    base = await _make_base(db_session, test_region)
    contract = await _make_contract(db_session, test_region)

    # Tour livraison avec contrat, 2 arrêts, valeurs numériques et temps chauffeur
    await _make_tour(
        db_session, base, code=f"H1-{uuid.uuid4().hex[:4]}", date="2026-06-01",
        contract=contract, n_stops=2, region=test_region,
        total_eqp=33, total_km=123.4, total_cost=456.78,
        departure_time="08:00", return_time="14:30", priority=1,
        driver_name="Jean Dupont",
        driver_arrival_time="2026-06-01T07:45",
    )

    resp = await client.get("/api/exports/tour-history")
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Historique tours"]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]

    assert list(rows[0]) == EXPECTED_HEADERS

    by_code = {r[0]: r for r in rows[1:]}
    r = next(v for k, v in by_code.items() if k.startswith("H1-"))
    assert r[1] == "Livraison"                       # Nature
    assert r[2] == "2026-06-01"                       # Date
    assert r[3] == base.name                          # Base (nom résolu)
    assert r[4].startswith(contract.vehicle_code)     # Véhicule résolu depuis contrat
    assert r[5] == "Transports Test"                  # Transporteur
    assert r[6] == 2                                   # Arrêts
    assert r[7] == 33                                  # EQC
    assert r[8] == pytest.approx(123.4)                # Km (float, pas Decimal)
    assert r[9] == pytest.approx(456.78)               # Coût
    assert r[10] == "Terminé"                          # Statut FR
    assert r[11] == "08:00"                            # Départ prévu
    assert r[12] == 1                                   # Priorité
    assert r[13] == "14:30"                             # Retour prévu
    assert r[14] == "Jean Dupont"                       # Chauffeur
    assert r[15] == "2026-06-01 07:45"                  # Arrivée chauffeur (T → espace)


@pytest.mark.asyncio
async def test_tour_history_export_no_contract_and_region_filter(client, db_session, test_region):
    from app.models.country import Country
    from app.models.region import Region

    base = await _make_base(db_session, test_region)
    await _make_tour(
        db_session, base, code=f"H2-{uuid.uuid4().hex[:4]}", date="2026-06-02",
        contract=None, n_stops=0, region=test_region,
    )

    # Tour dans une AUTRE région, doit être exclu par le filtre region_id
    country = (await db_session.execute(
        __import__("sqlalchemy").select(Country).where(Country.code == "TST")
    )).scalar_one()
    other_region = Region(name=f"Other {uuid.uuid4().hex[:5]}", country_id=country.id)
    db_session.add(other_region)
    await db_session.commit()
    await db_session.refresh(other_region)
    other_base = await _make_base(db_session, other_region, name="Base Autre")
    await _make_tour(
        db_session, other_base, code=f"HX-{uuid.uuid4().hex[:4]}", date="2026-06-03",
        contract=None, n_stops=0, region=other_region,
    )

    resp = await client.get("/api/exports/tour-history", params={"region_id": test_region.id})
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    codes = [r[0] for r in wb["Historique tours"].iter_rows(min_row=2, values_only=True)]

    assert any(c and c.startswith("H2-") for c in codes)
    assert not any(c and c.startswith("HX-") for c in codes)  # autre région exclue

    # Tour sans contrat : véhicule/transporteur vides
    # (openpyxl relit une cellule vide comme None au round-trip)
    r2 = next(r for r in wb["Historique tours"].iter_rows(min_row=2, values_only=True) if r[0].startswith("H2-"))
    assert r2[4] in (None, "") and r2[5] in (None, "")
