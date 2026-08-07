"""Test export planning postier — feuille « Tours » (Tournées ERT).

Vérifie la mise en page reproduite de 07-08 Planning.xlsm : titre, en-têtes
en ligne 6, une ligne par tournée (ordre ERT), séquence PDV+EQC, base
« code (name) », et colonnes d'exploitation mappées sur les champs Tour.
"""

import io
import uuid

import pytest
from openpyxl import load_workbook


async def _make_base(db_session, region, code="080", name="Villers"):
    from app.models.base_logistics import BaseLogistics

    base = BaseLogistics(code=f"{code}{uuid.uuid4().hex[:3]}", name=name, city="Villers", region_id=region.id)
    db_session.add(base)
    await db_session.commit()
    await db_session.refresh(base)
    return base


async def _make_pdv(db_session, region, code, city):
    from app.models.pdv import PDV, PDVType

    pdv = PDV(code=code, name=f"PDV {code}", type=PDVType.HYPER, city=city, region_id=region.id)
    db_session.add(pdv)
    await db_session.commit()
    await db_session.refresh(pdv)
    return pdv


async def _make_tour(db_session, base, region, *, code, priority, temp, departure, stops):
    from app.models.tour import Tour, TourStatus
    from app.models.tour_stop import TourStop

    tour = Tour(
        date="2026-08-07", delivery_date="2026-08-07", code=code, base_id=base.id,
        status=TourStatus.VALIDATED, priority=priority, temperature_type=temp,
        departure_time=departure, total_eqp=42, driver_name="Baeskens",
    )
    db_session.add(tour)
    await db_session.commit()
    await db_session.refresh(tour)
    for seq, (pcode, city, eqc) in enumerate(stops, start=1):
        pdv = await _make_pdv(db_session, region, pcode, city)
        db_session.add(TourStop(tour_id=tour.id, pdv_id=pdv.id, sequence_order=seq, eqp_count=eqc))
    await db_session.commit()
    return tour


@pytest.mark.asyncio
async def test_postier_planning_layout_and_order(client, db_session, test_region):
    base = await _make_base(db_session, test_region, code="2785", name="Trazegnies")

    # priorité 2 (doit sortir en 2e), 1 arrêt
    await _make_tour(db_session, base, test_region, code=f"B-{uuid.uuid4().hex[:4]}",
                     priority=2, temp="FRAIS", departure="06:00",
                     stops=[("6307", "Dison", 10)])
    # priorité 1 (doit sortir en 1er), 2 arrêts, GEL
    await _make_tour(db_session, base, test_region, code=f"A-{uuid.uuid4().hex[:4]}",
                     priority=1, temp="GEL", departure="05:00",
                     stops=[("6060", "Herve", 5), ("9879", "Heusy", 12)])

    resp = await client.get("/api/exports/postier-planning", params={"date": "2026-08-07", "base_id": base.id})
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Tours"]

    # Titre + en-têtes ligne 6
    assert ws.cell(2, 3).value == "Tournées ERT du"
    assert ws.cell(6, 1).value == "Ordre"
    assert ws.cell(6, 2).value == "N° Mission"
    assert ws.cell(6, 14).value == "PDV 1"
    assert ws.cell(6, 15).value == "E.P1"
    assert ws.cell(6, 46).value == "H.Départ"        # 1re colonne d'exploitation (AT)
    assert ws.cell(6, 50).value == "Eqc Prévis."

    # Données : ordre ERT → priorité 1 en premier (ligne 7)
    assert ws.cell(7, 1).value == 1                    # Ordre
    assert ws.cell(7, 9).value == "O"                  # Gel (GEL)
    assert ws.cell(7, 12).value == f"{base.code} (Trazegnies)"  # Départ = code (name)
    assert ws.cell(7, 14).value == "6060 (Herve)"      # PDV 1 = code (city)
    assert ws.cell(7, 15).value == 5                   # E.P1
    assert ws.cell(7, 16).value == "9879 (Heusy)"      # PDV 2
    assert ws.cell(7, 17).value == 12                  # E.P2
    assert ws.cell(7, 46).value == "05:00"             # H.Départ
    assert ws.cell(7, 50).value == 42                  # Eqc Prévis.

    # 2e ligne = priorité 2, non-GEL
    assert ws.cell(8, 1).value == 2
    assert ws.cell(8, 9).value == "N"
    assert ws.cell(8, 14).value == "6307 (Dison)"
