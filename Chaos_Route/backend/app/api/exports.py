"""Routes Export CSV/Excel / Export API routes."""

from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import io

from app.database import get_db
from app.models.country import Country
from app.models.region import Region
from app.models.base_logistics import BaseLogistics
from app.models.pdv import PDV
from app.models.parameter import Parameter
from app.models.supplier import Supplier
from app.models.tour import Tour
from app.models.tour_stop import TourStop
from app.models.volume import Volume
from app.models.contract import Contract
from app.models.distance_matrix import DistanceMatrix
from app.models.km_tax import KmTax
from app.models.cnuf_temperature import CnufTemperature
from app.models.user import User
from app.services.export_service import ExportService
from app.api.deps import require_permission, get_user_region_ids

router = APIRouter()

# Code transporteur Infolog par défaut (transport propre CMRO) si le paramètre
# wms_infolog_carrier_code n'est pas défini / Default Infolog carrier code.
DEFAULT_WMS_CARRIER_CODE = "08000888"

# Mapping entité -> modèle SQLAlchemy / Entity to model mapping
ENTITY_MODEL_MAP = {
    "countries": Country,
    "regions": Region,
    "bases": BaseLogistics,
    "pdvs": PDV,
    "suppliers": Supplier,
    "volumes": Volume,
    "contracts": Contract,
    "distances": DistanceMatrix,
    "km-tax": KmTax,
    "cnuf-temperatures": CnufTemperature,
}

# Entités avec filtrage par région / Entities with region scoping
REGION_SCOPED_ENTITIES = {"bases", "pdvs", "suppliers", "contracts"}


# NB : déclarée AVANT la route dynamique /{entity_type} pour ne pas être
# capturée par celle-ci / Declared BEFORE /{entity_type} so it isn't shadowed.
@router.get("/wms-infolog")
async def export_wms_infolog(
    date: str = Query(..., description="Date de planification YYYY-MM-DD"),
    base_id: int | None = Query(None, description="Filtrer sur une base logistique"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("tour-planning", "read")),
):
    """Export WMS Infolog (TMS_vers_wms).

    Génère le fichier Excel attendu par la macro d'encodage Infolog : une ligne
    par arrêt PDV, les tours rangés dans l'ordre ERT (priorité), et les PDV de
    chaque tour en ordre INVERSE (le dernier livré encodé en premier).

    Colonnes (A→H, sans en-tête, feuille « Export ») :
    A = ordre ERT (priorité) · B = code PDV · C = code chauffeur Infolog ·
    D = code transporteur · E = date de livraison · F = heure de départ ·
    G = index global (décroissant par tour) · H = heure de départ (texte).
    """
    # ── Code transporteur (paramètre global configurable) ────────────────────
    carrier_param = await db.execute(
        select(Parameter).where(
            Parameter.key == "wms_infolog_carrier_code",
            Parameter.region_id.is_(None),
        )
    )
    carrier_row = carrier_param.scalar_one_or_none()
    carrier_code = (carrier_row.value if carrier_row and carrier_row.value else DEFAULT_WMS_CARRIER_CODE)

    # ── Tours planifiés du jour (heure de départ renseignée) ─────────────────
    query = (
        select(Tour)
        .where(Tour.date == date, Tour.departure_time.isnot(None))
        .options(selectinload(Tour.stops).selectinload(TourStop.pdv))
    )
    if base_id is not None:
        query = query.where(Tour.base_id == base_id)
    else:
        region_ids = get_user_region_ids(user)
        if region_ids is not None:
            # Restreindre aux bases des régions de l'utilisateur / Scope to user's regions
            base_q = select(BaseLogistics.id).where(BaseLogistics.region_id.in_(region_ids))
            allowed_bases = (await db.execute(base_q)).scalars().all()
            query = query.where(Tour.base_id.in_(allowed_bases))

    tours = list((await db.execute(query)).scalars().all())

    # Ordre ERT : priorité (1..n) d'abord, NULL en dernier, puis heure puis code /
    # ERT order: priority first (NULLs last), then departure time, then code
    tours.sort(key=lambda t: (
        t.priority is None,
        t.priority if t.priority is not None else 0,
        t.departure_time or "",
        t.code,
    ))

    def _hhmmss(t: str | None) -> str:
        """HH:MM → HH:MM:SS (texte attendu par la macro)."""
        if not t:
            return ""
        parts = t.split(":")
        h = parts[0] if len(parts) > 0 else "00"
        m = parts[1] if len(parts) > 1 else "00"
        s = parts[2] if len(parts) > 2 else "00"
        return f"{int(h):02d}:{int(m):02d}:{int(s):02d}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    global_index = 0
    rank = 0
    for tour in tours:
        rank += 1
        ordre = tour.priority if tour.priority is not None else rank
        # Arrêts du tour dans l'ordre de livraison (1..n) /
        # Tour stops in delivery order (1..n)
        stops = sorted(tour.stops, key=lambda s: s.sequence_order)
        if not stops:
            continue
        n = len(stops)
        # Bloc d'index global contigu [global_index+1 .. global_index+n], affecté
        # dans l'ordre de livraison ; sortie en ordre inverse (index décroissant).
        base_index = global_index
        global_index += n

        delivery_date_str = tour.delivery_date or tour.date
        try:
            y, mo, d = (int(x) for x in delivery_date_str.split("-"))
            delivery_date_val: object = datetime(y, mo, d)
        except (ValueError, AttributeError):
            delivery_date_val = delivery_date_str

        dep_text = _hhmmss(tour.departure_time)

        # Inverser : le dernier PDV livré (sequence max) en premier /
        # Reverse: last delivered PDV (max sequence) first
        for pos, stop in enumerate(reversed(stops)):
            # rang de livraison 1..n (1 = premier livré) → index global croissant
            delivery_rank = n - pos  # le dernier livré a delivery_rank = n
            idx = base_index + delivery_rank
            pdv_code = stop.pdv.code if stop.pdv else ""
            ws.append([
                ordre,                              # A
                pdv_code,                           # B
                tour.driver_code_infolog or "",     # C
                carrier_code,                       # D
                delivery_date_val,                  # E
                dep_text,                           # F
                idx,                                # G
                dep_text,                           # H
            ])

    content = io.BytesIO()
    wb.save(content)
    content.seek(0)
    filename = f"TMS_vers_wms_{date}.xlsx"
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Libellés FR alignés sur l'historique web / FR labels mirroring the web history.
_TOUR_TYPE_LABELS = {
    "ENLEVEMENT": "Enlèvement",
    "VIDANGES": "Vidanges",
    "DEPLACEMENT_BASE": "Déplacement",
    "GARAGE": "Garage",
    "TRANSFERT_PDV": "Transfert PDV",
    "ENLEVEMENT_DEDIE": "Enlèvement dédié",
}
_TOUR_STATUS_LABELS = {
    "DRAFT": "Brouillon",
    "VALIDATED": "Validé",
    "IN_PROGRESS": "En cours",
    "RETURNING": "En retour",
    "COMPLETED": "Terminé",
}


def _dt_local(value: str | None) -> str:
    """Rend un datetime-local (YYYY-MM-DDTHH:MM) lisible (T → espace)."""
    if not value:
        return ""
    return value.replace("T", " ")


def _num(value: object) -> float | None:
    """Numeric/Decimal → float (openpyxl ne sérialise pas Decimal)."""
    return float(value) if value is not None else None


# NB : déclarée AVANT la route dynamique /{entity_type} pour ne pas être capturée
# par celle-ci / Declared BEFORE /{entity_type} so it isn't shadowed.
@router.get("/tour-history")
async def export_tour_history(
    region_id: int | None = Query(None, description="Filtrer sur une région"),
    base_id: int | None = Query(None, description="Filtrer sur une base logistique"),
    date_from: str | None = Query(None, description="Date de début (YYYY-MM-DD, incluse)"),
    date_to: str | None = Query(None, description="Date de fin (YYYY-MM-DD, incluse)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("tour-history", "read")),
):
    """Export Excel de l'historique des tours (ticket #17, phase 1 — manuel).

    Reprend les colonnes complètes de la page « Historique des tours ». Une ligne
    par tour, respectant le périmètre régional de l'utilisateur et les filtres
    optionnels (région, base, plage de dates).
    """
    query = (
        select(Tour)
        .options(
            selectinload(Tour.stops),
            selectinload(Tour.contract),
            selectinload(Tour.base),
        )
    )
    if base_id is not None:
        query = query.where(Tour.base_id == base_id)
    if date_from is not None:
        query = query.where(Tour.date >= date_from)
    if date_to is not None:
        query = query.where(Tour.date <= date_to)

    # Périmètre région : filtre demandé + périmètre de l'utilisateur /
    # Region scope: requested filter + user's own scope
    region_filter_ids: list[int] | None = None
    user_region_ids = get_user_region_ids(user)
    if region_id is not None:
        region_filter_ids = [region_id]
    if user_region_ids is not None:
        region_filter_ids = (
            [r for r in region_filter_ids if r in user_region_ids]
            if region_filter_ids is not None else list(user_region_ids)
        )
    if region_filter_ids is not None:
        base_q = select(BaseLogistics.id).where(BaseLogistics.region_id.in_(region_filter_ids))
        allowed_bases = (await db.execute(base_q)).scalars().all()
        query = query.where(Tour.base_id.in_(allowed_bases))

    query = query.order_by(Tour.date.desc(), Tour.id.desc())
    tours = list((await db.execute(query)).scalars().all())

    headers = [
        "Code", "Nature", "Date", "Base", "Véhicule", "Transporteur", "Arrêts",
        "EQC", "Km", "Coût (€)", "Statut", "Départ prévu", "Priorité",
        "Retour prévu", "Chauffeur", "Arrivée chauffeur", "Fin chargement",
        "Top départ", "Sortie barrière", "Retour barrière",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Historique tours"
    ws.append(headers)

    for tour in tours:
        ttype = tour.tour_type.value if tour.tour_type is not None else "LIVRAISON"
        nature = "Livraison" if ttype == "LIVRAISON" else _TOUR_TYPE_LABELS.get(ttype, ttype)

        c = tour.contract
        if c is None:
            vehicle = ""
            transporter = ""
        else:
            vehicle = f"{c.vehicle_code} — {c.vehicle_name or ''}" if c.vehicle_code else c.code
            transporter = c.transporter_name or ""

        ws.append([
            tour.code,
            nature,
            tour.date,
            tour.base.name if tour.base is not None else f"#{tour.base_id}",
            vehicle,
            transporter,
            len(tour.stops),
            _num(tour.total_eqp),
            _num(tour.total_km),
            _num(tour.total_cost),
            _TOUR_STATUS_LABELS.get(tour.status.value, tour.status.value),
            tour.departure_time or "",
            tour.priority if tour.priority is not None else "",
            tour.return_time or "",
            tour.driver_name or "",
            _dt_local(tour.driver_arrival_time),
            _dt_local(tour.loading_end_time),
            _dt_local(tour.departure_signal_time),
            _dt_local(tour.barrier_exit_time),
            _dt_local(tour.barrier_entry_time),
        ])

    # En-tête en gras + volet figé / Bold header + freeze pane
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    content = io.BytesIO()
    wb.save(content)
    content.seek(0)
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"historique-tours_{today}.xlsx"
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Colonnes de la feuille « Tours » (Tournées ERT) de 07-08 Planning.xlsm reproduites
# à l'identique : infos tournée (A-M), 16 paires PDV/EQC (N-AS), colonnes
# d'exploitation postier/garde (AT-BG). / "Tours" (ERT) sheet columns reproduced.
_PLANNING_PDV_PAIRS = 16
_PLANNING_HEADERS = (
    ["Ordre", "N° Mission", "Chargeurs", "Code Ch.", "Type Tour", "Chauffeurs",
     "Trac", "Semi", "Gel", "TKT", "Observations/Enlèvement", "Départ", "Retour"]
    + [lab for k in range(1, _PLANNING_PDV_PAIRS + 1) for lab in (f"PDV {k}", f"E.P{k}")]
    + ["H.Départ", "Porte", "T°", "H. disp. Semi", "Eqc Prévis.", "EQC Chargés",
       "Top Départ", "Prés. sur site", "H.Sortie", "H.Retour", "Kms départ",
       "Kms retour", "KM calculé", "Remarque Garde"]
)
# Index 1-based de la 1re colonne PDV (N) et de la 1re colonne d'exploitation (AT)
_PLANNING_PDV_COL0 = 14
_PLANNING_OPS_COL0 = _PLANNING_PDV_COL0 + 2 * _PLANNING_PDV_PAIRS  # = 46 (AT)


@router.get("/postier-planning")
async def export_postier_planning(
    date: str = Query(..., description="Date de livraison (YYYY-MM-DD)"),
    base_id: int = Query(..., description="Base logistique"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("operations", "read")),
):
    """Export Excel du planning postier — feuille « Tours » (Tournées ERT).

    Reproduit la mise en page de 07-08 Planning.xlsm (feuille Tours) : une ligne
    par tournée du jour (ordre ERT), la séquence des PDV avec leur EQC, et les
    colonnes d'exploitation (tops, kms…) que le postier/la garde renseignent.
    Les colonnes sans équivalent dans l'app (Type Tour, TKT, Remarque Garde)
    restent vides.
    """
    query = (
        select(Tour)
        .where(
            Tour.base_id == base_id,
            or_(
                Tour.delivery_date == date,
                and_(Tour.delivery_date.is_(None), Tour.date == date),
            ),
        )
        .options(
            selectinload(Tour.stops).selectinload(TourStop.pdv),
            selectinload(Tour.base),
            selectinload(Tour.contract),
            selectinload(Tour.vehicle),
            selectinload(Tour.tractor),
        )
    )
    # Périmètre région de l'utilisateur / User region scope
    region_ids = get_user_region_ids(user)
    if region_ids is not None:
        allowed = (await db.execute(
            select(BaseLogistics.id).where(BaseLogistics.region_id.in_(region_ids))
        )).scalars().all()
        query = query.where(Tour.base_id.in_(allowed))

    tours = list((await db.execute(query)).scalars().all())
    # Ordre ERT : priorité (NULL en dernier), puis heure de départ, puis code
    tours.sort(key=lambda t: (
        t.priority is None, t.priority or 0, t.departure_time or "", t.code,
    ))

    def _base_disp(b) -> str:
        return f"{b.code} ({b.name})" if b else ""

    def _eqc(v) -> float | None:
        return float(v) if v is not None else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Tours"
    # Titre (comme l'original : « Tournées ERT du » + date)
    ws.cell(2, 3, "Tournées ERT du")
    ws.cell(2, 8, date)
    # En-têtes en ligne 6
    for i, label in enumerate(_PLANNING_HEADERS, start=1):
        cell = ws.cell(6, i, label)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A7"

    r = 7
    for rank, tour in enumerate(tours, start=1):
        base_disp = _base_disp(tour.base)
        gel = "O" if (tour.temperature_type and "GEL" in tour.temperature_type) else "N"
        trac = tour.tractor.code if tour.tractor else ""
        semi = (
            tour.vehicle.code if tour.vehicle
            else (tour.contract.vehicle_code if tour.contract and tour.contract.vehicle_code else (tour.trailer_number or ""))
        )
        ws.cell(r, 1, tour.priority if tour.priority is not None else rank)
        ws.cell(r, 2, tour.wms_tour_code or tour.code)
        ws.cell(r, 3, tour.loader_name or "")
        ws.cell(r, 4, tour.loader_code or "")
        # 5 = Type Tour (pas d'équivalent) — laissé vide
        ws.cell(r, 6, tour.driver_name or "")
        ws.cell(r, 7, trac)
        ws.cell(r, 8, semi)
        ws.cell(r, 9, gel)
        # 10 = TKT — laissé vide
        ws.cell(r, 11, tour.remarks or tour.destination or "")
        ws.cell(r, 12, base_disp)
        ws.cell(r, 13, base_disp)
        # PDV 1..16 + EQC (colonnes N..AS)
        stops = sorted(tour.stops, key=lambda s: s.sequence_order)
        for k, stop in enumerate(stops[:_PLANNING_PDV_PAIRS]):
            pdv = stop.pdv
            pdv_disp = (f"{pdv.code} ({pdv.city})" if pdv and pdv.city else (pdv.code if pdv else "")) if pdv else ""
            ws.cell(r, _PLANNING_PDV_COL0 + 2 * k, pdv_disp)
            ws.cell(r, _PLANNING_PDV_COL0 + 2 * k + 1, _eqc(stop.eqp_count))
        # Colonnes d'exploitation (AT..BG)
        c = _PLANNING_OPS_COL0
        ws.cell(r, c, tour.departure_time or "")               # AT H.Départ
        ws.cell(r, c + 1, tour.dock_door_number or "")          # AU Porte
        ws.cell(r, c + 2, tour.temperature_type or "")          # AV T°
        ws.cell(r, c + 3, _dt_local(tour.trailer_ready_time))   # AW H. disp. Semi
        ws.cell(r, c + 4, _eqc(tour.total_eqp))                 # AX Eqc Prévis.
        ws.cell(r, c + 5, tour.eqp_loaded)                      # AY EQC Chargés
        ws.cell(r, c + 6, _dt_local(tour.departure_signal_time))  # AZ Top Départ
        ws.cell(r, c + 7, _dt_local(tour.driver_arrival_time))  # BA Prés. sur site
        ws.cell(r, c + 8, _dt_local(tour.barrier_exit_time))    # BB H.Sortie
        ws.cell(r, c + 9, _dt_local(tour.barrier_entry_time))   # BC H.Retour
        ws.cell(r, c + 10, tour.km_departure)                   # BD Kms départ
        ws.cell(r, c + 11, tour.km_return)                      # BE Kms retour
        ws.cell(r, c + 12, _eqc(tour.total_km))                # BF KM calculé
        # BG Remarque Garde — laissé vide
        r += 1

    content = io.BytesIO()
    wb.save(content)
    content.seek(0)
    filename = f"planning-postier_{date}.xlsx"
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{entity_type}")
async def export_data(
    entity_type: str,
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("imports-exports", "read")),
):
    """Exporter les données d'une entité / Export entity data to CSV or XLSX."""
    if entity_type not in ENTITY_MODEL_MAP:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid entity type. Allowed: {list(ENTITY_MODEL_MAP.keys())}",
        )

    model_class = ENTITY_MODEL_MAP[entity_type]
    fields = ExportService.get_fields(entity_type)
    if not fields:
        raise HTTPException(status_code=400, detail=f"No field mapping for entity: {entity_type}")

    # Requête avec filtrage région si applicable / Query with region scoping if applicable
    query = select(model_class)
    if entity_type in REGION_SCOPED_ENTITIES:
        region_ids = get_user_region_ids(user)
        if region_ids is not None:
            query = query.where(model_class.region_id.in_(region_ids))

    result = await db.execute(query)
    objects = result.scalars().all()

    rows = [ExportService.model_to_dict(obj, fields) for obj in objects]

    # Remplacer DB IDs par codes pour import round-trip / Replace DB IDs with codes
    if entity_type in ("distances", "km-tax", "volumes"):
        pdv_id_to_code: dict[int, str] = {}
        base_id_to_code: dict[int, str] = {}
        r = await db.execute(select(PDV.id, PDV.code))
        for eid, code in r.all():
            pdv_id_to_code[eid] = str(code)
        r = await db.execute(select(BaseLogistics.id, BaseLogistics.code))
        for eid, code in r.all():
            base_id_to_code[eid] = str(code)

        if entity_type in ("distances", "km-tax"):
            sup_id_to_code: dict[int, str] = {}
            r = await db.execute(select(Supplier.id, Supplier.code))
            for eid, code in r.all():
                sup_id_to_code[eid] = str(code)
            type_lookup = {"PDV": pdv_id_to_code, "BASE": base_id_to_code, "SUPPLIER": sup_id_to_code}
            for row in rows:
                for prefix in ("origin", "destination"):
                    etype = row.get(f"{prefix}_type")
                    eid = row.get(f"{prefix}_id")
                    if etype and eid is not None:
                        row[f"{prefix}_id"] = type_lookup.get(etype, {}).get(eid, eid)
        elif entity_type == "volumes":
            for row in rows:
                pid = row.get("pdv_id")
                if pid is not None:
                    row["pdv_id"] = pdv_id_to_code.get(pid, pid)
                bid = row.get("base_origin_id")
                if bid is not None:
                    row["base_origin_id"] = base_id_to_code.get(bid, bid)

    if format == "csv":
        content = ExportService.to_csv(rows, fields)
        media_type = "text/csv; charset=utf-8"
        filename = f"{entity_type}.csv"
    else:
        content = ExportService.to_xlsx(rows, fields, sheet_name=entity_type.capitalize())
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{entity_type}.xlsx"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/tours/{tour_id}/excel")
async def export_tour_excel(
    tour_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("imports-exports", "read")),
):
    """Exporter un tour en Excel / Export a tour to Excel."""
    # TODO: Implémenter l'export Excel / Implement Excel export
    return {"status": "pending", "message": "Excel export will be implemented in Phase 4"}


@router.get("/tours/{tour_id}/pdf")
async def export_tour_pdf(
    tour_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_permission("imports-exports", "read")),
):
    """Exporter un tour en PDF / Export a tour to PDF."""
    # TODO: Implémenter l'export PDF / Implement PDF export
    return {"status": "pending", "message": "PDF export will be implemented in Phase 4"}
